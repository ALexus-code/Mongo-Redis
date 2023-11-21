from openpyxl import load_workbook
import uuid
import jinja2


environment = jinja2.Environment()
shablon = environment.from_string("'{{name}}' : ' {{str_}} ',")


#выводит все
# for row in sheet.iter_rows():
#     for cell in row:
#         print(cell.value, end=' ')
#     print()
# for column in sheet:
#     print(column[1])

#парсинг для файла
def xsls2jsonstrings(file_name):

    book = load_workbook(filename=str(file_name))
    sheet = book.active
    stroki = []
    uuid4s = []
    for row in range(2, sheet.max_row):
        Nazvanie = shablon.render(name='Название', str_=str(sheet[row][0].value).replace('"', ''))
        INN = shablon.render(name='ИНН', str_=str(sheet[row][1].value).replace('"', ''))
        KPP = shablon.render(name='КПП', str_=str(sheet[row][2].value).replace('"', ''))
        telephone = shablon.render(name='Телефон', str_=str(sheet[row][3].value).replace('"', ''))
        www = shablon.render(name='www', str_=str(sheet[row][4].value).replace('"', ''))
        ONLAINPLATEZH = shablon.render(name='НеобходимостьФискализацииПлатежейОнлайнКасса',str_=str(sheet[row][0].value).replace('"', ''))
        GIS_DATA = shablon.render(name='ГИС_ДАТА', str_=str(sheet[row][5].value).replace('"', ''))
        GIS_GUID = shablon.render(name='ГИС_ГУИД', str_=str(sheet[row][6].value).replace('"', ''))
        RFIOR = shablon.render(name='РФИОРуководителя', str_=str(sheet[row][7].value).replace('"', ''))
        IDD = shablon.render(name='ИДДляДокументооборота', str_=str(sheet[row][8].value).replace('"', ''))
        telephone_cont = shablon.render(name='ТелефонКонт', str_=str(sheet[row][9].value).replace('"', ''))
        type_user = shablon.render(name='ТипПользователяВеб', str_=str(sheet[row][10].value).replace('"', ''))
        licvidation_data = shablon.render(name='ДатаЛиквидации', str_=str(sheet[row][11].value).replace('"', ''))
        registration_data = shablon.render(name='ДатаРегистрации', str_=str(sheet[row][12].value).replace('"', ''))
        category = shablon.render(name='Категория', str_=str(sheet[row][13].value).replace('"', ''))
        Fact_adres_kladr = shablon.render(name='ФактАдресКладр', str_=str(sheet[row][14].value).replace('"', ''))
        Adres_kladr = shablon.render(name='АдресКладр', str_=str(sheet[row][15].value).replace('"', ''))
        adres_detail = shablon.render(name='АдресДетально', str_=str(sheet[row][16].value).replace('"', ''))
        fact_adres = shablon.render(name='ФактАдрес', str_=str(sheet[row][17].value).replace('"', ''))
        date_OGRN = shablon.render(name='ДатаОГРН', str_=str(sheet[row][18].value).replace('"', ''))
        PP_group = shablon.render(name='ГруппировкаПП', str_=str(sheet[row][19].value).replace('"', ''))
        zareg_org = shablon.render(name='ЗарегистрировавшийОрган', str_=str(sheet[row][20].value).replace('"', ''))
        facs = shablon.render(name='Факс', str_=str(sheet[row][21].value).replace('"', ''))
        OGRN = shablon.render(name='ОГРН', str_=str(sheet[row][22].value).replace('"', ''))
        chasi_raboti_pasportnogo = shablon.render(name='Часы работы паспортного',str_=str(sheet[row][23].value).replace('"', ''))
        chasi_raboti_kassi = shablon.render(name='Часы работы кассы', str_=str(sheet[row][24].value).replace('"', ''))
        chasi_raboty_buhg = shablon.render(name='Часы работы бухгалтерии',str_=str(sheet[row][25].value).replace('"', ''))
        FIO_rukovod = shablon.render(name='ФИОРуководителя', str_=str(sheet[row][26].value).replace('"', ''))
        FIO_main_buhg = shablon.render(name='ФИОГлБухгалтера', str_=str(sheet[row][27].value).replace('"', ''))
        massege = shablon.render(name='Сообщение', str_=str(sheet[row][28].value).replace('"', ''))
        Rezhim_raboti = shablon.render(name='Режим работы', str_=str(sheet[row][29].value).replace('"', ''))
        primechanie = shablon.render(name='Примечание', str_=str(sheet[row][30].value).replace('"', ''))
        Otrasl = shablon.render(name='Отрасль', str_=str(sheet[row][31].value).replace('"', ''))
        OtdelKKM = shablon.render(name='ОтделККМ', str_=str(sheet[row][32].value).replace('"', ''))
        OKDP = shablon.render(name='ОКДП', str_=str(sheet[row][33].value).replace('"', ''))
        OKATO = shablon.render(name='ОКАТО', str_=str(sheet[row][34].value).replace('"', ''))
        OKPO = shablon.render(name='ОКПО', str_=str(sheet[row][35].value).replace('"', ''))
        OKONH = shablon.render(name='ОКОНХ', str_=str(sheet[row][36].value).replace('"', ''))
        OKVED = shablon.render(name='ОКВЭД', str_=str(sheet[row][37].value).replace('"', ''))
        kom_sbor = shablon.render(name='КомСбор', str_=str(sheet[row][38].value).replace('"', ''))
        dostup = shablon.render(name='Доступ', str_=str(sheet[row][39].value).replace('"', ''))
        Tnaimen = shablon.render(name='ТНаименование', str_=str(sheet[row][40].value).replace('"', ''))
        Rnaimen = shablon.render(name='РНаименование', str_=str(sheet[row][41].value).replace('"', ''))
        Dnaimen = shablon.render(name='ДНаименование', str_=str(sheet[row][42].value).replace('"', ''))
        naimen = shablon.render(name='Наименование', str_=str(sheet[row][43].value).replace('"', ''))
        Dolzhnost_rucovod = shablon.render(name='ДолжностьРуководителя', str_=str(sheet[row][44].value).replace('"', ''))
        stavka_NDS = shablon.render(name='СтавкаНДС', str_=str(sheet[row][45].value).replace('"', ''))
        variant_NDS = shablon.render(name='Вариант НДС', str_=str(sheet[row][46].value).replace('"', ''))
        budzhet = shablon.render(name='Бюджет', str_=str(sheet[row][47].value).replace('"', ''))
        email = shablon.render(name='email', str_=str(sheet[row][48].value).replace('"', ''))
        energettel = shablon.render(name='ЭнергетТел', str_=str(sheet[row][49].value).replace('"', ''))
        energetemail = shablon.render(name='ЭнергетEmail', str_=str(sheet[row][50].value).replace('"', ''))
        Buhtel = shablon.render(name='БухТел', str_=str(sheet[row][51].value).replace('"', ''))
        Buhemail = shablon.render(name='БухEmail', str_=str(sheet[row][52].value).replace('"', ''))
        Ruktel = shablon.render(name='РукТел', str_=str(sheet[row][53].value).replace('"', ''))
        Rukemail = shablon.render(name='РукEmail', str_=str(sheet[row][54].value).replace('"', ''))
        adres = shablon.render(name='Адрес', str_=str(sheet[row][55].value).replace('"', ''))[:-1]
        new_id = str(uuid.uuid4())
        id = '"uuid4":"' + new_id + '",'
        #new_str = '"' + new_id + '"' + " = {" + Nazvanie + INN + KPP + telephone + www + ONLAINPLATEZH + GIS_DATA + GIS_GUID + RFIOR + IDD + telephone_cont + type_user + licvidation_data + registration_data + category + Fact_adres_kladr + Adres_kladr + adres_detail + fact_adres + date_OGRN + PP_group + zareg_org + facs + OGRN + chasi_raboti_pasportnogo + chasi_raboti_kassi + chasi_raboty_buhg + FIO_rukovod + FIO_main_buhg + massege + Rezhim_raboti + primechanie + Otrasl + OtdelKKM + OKDP + OKONH + OKVED + kom_sbor + dostup + Tnaimen + Rnaimen + Dnaimen + naimen + Dolzhnost_rucovod + stavka_NDS + variant_NDS + budzhet + email + energettel + energetemail + Buhtel + Buhemail + Ruktel + Rukemail + adres + '}'
        full_json = '{' + id + Nazvanie + INN + KPP + telephone + www + ONLAINPLATEZH + GIS_DATA + GIS_GUID + RFIOR + IDD + telephone_cont + type_user + licvidation_data + registration_data + category + Fact_adres_kladr + Adres_kladr + adres_detail + fact_adres + date_OGRN + PP_group + zareg_org + facs + OGRN + chasi_raboti_pasportnogo + chasi_raboti_kassi + chasi_raboty_buhg + FIO_rukovod + FIO_main_buhg + massege + Rezhim_raboti + primechanie + Otrasl + OtdelKKM + OKDP + OKONH + OKVED + kom_sbor + dostup + Tnaimen + Rnaimen + Dnaimen + naimen + Dolzhnost_rucovod + stavka_NDS + variant_NDS + budzhet + email + energettel + energetemail + Buhtel + Buhemail + Ruktel + Rukemail + adres + '}'#]}'
        #'{"enterprise' + str(row) + '": [
        import json
        new_str = full_json.replace("'",'"')
        stroki.append(new_str)
        uuid4s.append(new_id)
        # print(new_id)
        # print(str(row))
    # print(stroki[1])
    return(stroki)
#print(uuid4s)
#print(xsls2jsonstrings('стек юл организации.xlsx'))